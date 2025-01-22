import openpyxl
from openpyxl.utils import rows_from_range

import Utilities
path=r'C:\Users\QSPR\Documents\DDT_data.xlsx'

#writing data
# workbook=openpyxl.load_workbook(path)
# sheet=workbook.active     #or  workbook["sheetname"]
# for r in range(1,4+1):
#     for c in range(1,2+1):
#         sheet.cell(r,c).value=input('enter the data:')
# workbook.save(path)

# #reading data
# workbook=openpyxl.load_workbook(path)
# sheet=workbook.active
# rows=sheet.max_row
# columns=sheet.max_column
# for r in range(1,rows+1):
#     for c in range(1,columns+1):
#         print(sheet.cell(r,c).value,end=' ')
#     print()

#write_data
# for r in range(1,4+1):
#     for c in range(1,2+1):
#         Utilities.write_data(path,"Sheet1",r,c,input('enter data:'))

#read_data
rows=Utilities.row_count(path,"Sheet1")
columns=Utilities.column_count(path,"Sheet1")

for r in range(1,rows+1):
    for c in range(1,columns+1):
        print(Utilities.read_data(path,"Sheet1",r,c),end=' ')
    print()